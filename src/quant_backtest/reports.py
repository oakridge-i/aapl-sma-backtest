from __future__ import annotations

import json
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .data import save_price_snapshot
from .experiments import ResearchResult


CSV_OUTPUTS = {
    "base_backtest.csv": "base_backtest",
    "cost_sensitivity.csv": "cost_sensitivity",
    "parameter_sweep.csv": "parameter_sweep",
    "train_test_results.csv": "train_test_results",
    "walk_forward_results.csv": "walk_forward_results",
    "multi_asset_results.csv": "multi_asset_results",
    "model_leaderboard.csv": "model_leaderboard",
    "hysteresis_sweep.csv": "hysteresis_sweep",
    "allocation_leaderboard.csv": "allocation_leaderboard",
    "capture_analysis.csv": "capture_analysis",
    "turnover_analysis.csv": "turnover_analysis",
    "v03_comparison.csv": "v03_comparison",
    "v03_cost_sensitivity.csv": "v03_cost_sensitivity",
    "capture_leaderboard.csv": "capture_leaderboard",
    "risk_filter_sweep.csv": "risk_filter_sweep",
    "regime_results.csv": "regime_results",
    "trade_log.csv": "trade_log",
    "benchmark_comparison.csv": "benchmark_comparison",
    "v04_comparison.csv": "v04_comparison",
    "v04_cost_sensitivity.csv": "v04_cost_sensitivity",
    "final_model_walk_forward.csv": "final_walk_forward",
    "significance_results.csv": "significance_results",
    "nested_walk_forward.csv": "nested_walk_forward",
    "nested_walk_forward_summary.csv": "nested_walk_forward_summary",
    "pbo_results.csv": "pbo_results",
    "family_leaderboard.csv": "family_leaderboard",
    "ensemble_leaderboard.csv": "ensemble_leaderboard",
    "overlay_leaderboard.csv": "overlay_leaderboard",
    "v06_comparison.csv": "v06_comparison",
    "v06_cost_sensitivity.csv": "v06_cost_sensitivity",
    "nested_ensemble_walk_forward.csv": "nested_ensemble_walk_forward",
    "nested_ensemble_summary.csv": "nested_ensemble_summary",
}


def save_research_outputs(result: ResearchResult, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    result.baseline_curve.to_csv(output_dir / "base_backtest.csv", index_label="Date")
    result.cost_sensitivity.to_csv(output_dir / "cost_sensitivity.csv", index=False)
    result.parameter_sweep.to_csv(output_dir / "parameter_sweep.csv", index=False)
    result.train_test_results.to_csv(output_dir / "train_test_results.csv", index=False)
    result.walk_forward_results.to_csv(output_dir / "walk_forward_results.csv", index=False)
    result.multi_asset_results.to_csv(output_dir / "multi_asset_results.csv", index=False)
    result.model_leaderboard.to_csv(output_dir / "model_leaderboard.csv", index=False)
    result.hysteresis_sweep.to_csv(output_dir / "hysteresis_sweep.csv", index=False)
    result.allocation_leaderboard.to_csv(output_dir / "allocation_leaderboard.csv", index=False)
    result.capture_analysis.to_csv(output_dir / "capture_analysis.csv", index=False)
    result.turnover_analysis.to_csv(output_dir / "turnover_analysis.csv", index=False)
    result.v03_comparison.to_csv(output_dir / "v03_comparison.csv", index=False)
    result.v03_cost_sensitivity.to_csv(output_dir / "v03_cost_sensitivity.csv", index=False)
    result.v03_curve.to_csv(output_dir / "v03_selected_curve.csv", index_label="Date")
    result.capture_leaderboard.to_csv(output_dir / "capture_leaderboard.csv", index=False)
    result.risk_filter_sweep.to_csv(output_dir / "risk_filter_sweep.csv", index=False)
    result.regime_results.to_csv(output_dir / "regime_results.csv", index=False)
    result.trade_log.to_csv(output_dir / "trade_log.csv", index=False)
    result.benchmark_comparison.to_csv(output_dir / "benchmark_comparison.csv", index=False)
    result.v04_comparison.to_csv(output_dir / "v04_comparison.csv", index=False)
    result.v04_cost_sensitivity.to_csv(output_dir / "v04_cost_sensitivity.csv", index=False)
    result.v04_curve.to_csv(output_dir / "v04_selected_curve.csv", index_label="Date")
    result.final_walk_forward.to_csv(output_dir / "final_model_walk_forward.csv", index=False)
    result.significance_results.to_csv(output_dir / "significance_results.csv", index=False)
    result.nested_walk_forward.to_csv(output_dir / "nested_walk_forward.csv", index=False)
    result.nested_walk_forward_summary.to_csv(output_dir / "nested_walk_forward_summary.csv", index=False)
    result.pbo_results.to_csv(output_dir / "pbo_results.csv", index=False)
    result.family_leaderboard.to_csv(output_dir / "family_leaderboard.csv", index=False)
    result.ensemble_leaderboard.to_csv(output_dir / "ensemble_leaderboard.csv", index=False)
    result.overlay_leaderboard.to_csv(output_dir / "overlay_leaderboard.csv", index=False)
    result.v06_comparison.to_csv(output_dir / "v06_comparison.csv", index=False)
    result.v06_cost_sensitivity.to_csv(output_dir / "v06_cost_sensitivity.csv", index=False)
    result.v06_curve.to_csv(output_dir / "v06_selected_curve.csv", index_label="Date")
    result.nested_ensemble_walk_forward.to_csv(output_dir / "nested_ensemble_walk_forward.csv", index=False)
    result.nested_ensemble_summary.to_csv(output_dir / "nested_ensemble_summary.csv", index=False)

    save_price_snapshot(result.prices, output_dir / "data_snapshot.csv")
    if result.run_metadata:
        with (output_dir / "run_manifest.json").open("w", encoding="utf-8") as file:
            json.dump(result.run_metadata, file, indent=2, default=str)

    save_research_plots(result, output_dir)
    save_research_workbook(result, output_dir / "research_report.xlsx")


def save_research_plots(result: ResearchResult, output_dir: Path) -> None:
    _plot_baseline(result.baseline_curve, output_dir / "baseline_equity_drawdown.png")
    _plot_cost_sensitivity(result.cost_sensitivity, output_dir / "cost_sensitivity.png")
    _plot_heatmap(
        result.parameter_sweep,
        value_col="sharpe",
        title="SMA Parameter Sweep: Sharpe",
        output_path=output_dir / "sharpe_heatmap.png",
    )
    _plot_heatmap(
        result.parameter_sweep,
        value_col="cagr",
        title="SMA Parameter Sweep: CAGR",
        output_path=output_dir / "cagr_heatmap.png",
    )
    _plot_train_test(result.train_test_results, output_dir / "train_test.png")
    _plot_multi_asset(result.multi_asset_results, output_dir / "multi_asset_comparison.png")
    _plot_leaderboard(result.model_leaderboard, output_dir / "leaderboard_top_models.png")
    _plot_v03_equity_drawdown(result.v03_curve, output_dir / "v03_equity_drawdown.png")
    _plot_turnover_scatter(result.turnover_analysis, output_dir / "turnover_vs_sharpe.png")
    _plot_capture(result.capture_analysis, output_dir / "capture_ratio.png")
    _plot_allocation_exposure(result.v03_curve, output_dir / "allocation_exposure.png")
    _plot_v03_cost_sensitivity(result.v03_cost_sensitivity, output_dir / "v03_cost_sensitivity.png")
    _plot_v03_entry_exit_signals(result.v03_curve, output_dir / "v03_entry_exit_signals.png")
    _plot_v04_equity_drawdown(result.v04_curve, output_dir / "v04_equity_drawdown.png")
    _plot_v04_entry_exit_signals(result.v04_curve, output_dir / "v04_entry_exit_signals.png")
    _plot_v04_capture(result.v04_comparison, output_dir / "v04_capture_chart.png")
    _plot_regime_performance(result.regime_results, output_dir / "regime_performance.png")
    _plot_turnover_capture_spread(result.capture_leaderboard, output_dir / "turnover_capture_spread.png")
    _plot_exposure_sizing(result.v04_curve, output_dir / "exposure_sizing.png")
    _plot_final_walk_forward(result.final_walk_forward, output_dir / "final_model_walk_forward.png")
    _plot_v06_comparison(result.v06_comparison, output_dir / "v06_comparison.png")
    _plot_allocation_exposure(result.v06_curve, output_dir / "v06_exposure.png")


def save_research_workbook(result: ResearchResult, output_path: Path) -> None:
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    sheets = {
        "Baseline": result.base_backtest,
        "Cost Sensitivity": result.cost_sensitivity,
        "Parameter Sweep": result.parameter_sweep,
        "Train Test": result.train_test_results,
        "Walk Forward": result.walk_forward_results,
        "Multi Asset": result.multi_asset_results,
        "Model Leaderboard": result.model_leaderboard,
        "Hysteresis Sweep": result.hysteresis_sweep,
        "Allocation Leaderboard": result.allocation_leaderboard,
        "Capture Analysis": result.capture_analysis,
        "Turnover Analysis": result.turnover_analysis,
        "v0.3 Comparison": result.v03_comparison,
        "v0.3 Costs": result.v03_cost_sensitivity,
        "Capture Leaderboard": result.capture_leaderboard,
        "Risk Filter Sweep": result.risk_filter_sweep,
        "Regime Results": result.regime_results,
        "Trade Log": result.trade_log,
        "Benchmark Comparison": result.benchmark_comparison,
        "v0.4 Comparison": result.v04_comparison,
        "v0.4 Costs": result.v04_cost_sensitivity,
        "Final Walk Forward": result.final_walk_forward,
        "Significance": result.significance_results,
        "Nested Walk Forward": result.nested_walk_forward,
        "Nested WF Summary": result.nested_walk_forward_summary,
        "PBO": result.pbo_results,
        "Family Leaderboard": result.family_leaderboard,
        "Ensemble Leaderboard": result.ensemble_leaderboard,
        "Overlay Leaderboard": result.overlay_leaderboard,
        "v0.6 Comparison": result.v06_comparison,
        "v0.6 Costs": result.v06_cost_sensitivity,
        "Nested Ensemble WF": result.nested_ensemble_walk_forward,
        "Nested Ensemble Summary": result.nested_ensemble_summary,
        "Raw Results": _raw_results(result),
    }

    _write_dashboard(dashboard, result)
    for sheet_name, table in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        _write_table(sheet, table)

    workbook.save(output_path)


def _plot_baseline(curve: pd.DataFrame, output_path: Path) -> None:
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    axes[0].plot(curve.index, curve["strategy_equity"], label="strategy")
    axes[0].plot(curve.index, curve["buy_hold_equity"], label="buy and hold")
    axes[0].set_title("Baseline equity")
    axes[0].grid(alpha=0.25)
    axes[0].legend()
    axes[1].plot(curve.index, curve["strategy_drawdown"], label="strategy drawdown")
    axes[1].plot(curve.index, curve["buy_hold_drawdown"], label="buy and hold drawdown")
    axes[1].set_title("Drawdown")
    axes[1].grid(alpha=0.25)
    axes[1].legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_cost_sensitivity(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(10, 6))
    for label, group in table.groupby("label"):
        ordered = group.sort_values("cost_bps")
        axis.plot(ordered["cost_bps"], ordered["cagr"], marker="o", label=f"{label} CAGR")
        axis.plot(ordered["cost_bps"], ordered["sharpe"], marker="s", linestyle="--", label=f"{label} Sharpe")
    axis.set_title("Transaction cost sensitivity")
    axis.set_xlabel("Cost assumption, bps")
    axis.grid(alpha=0.25)
    axis.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_heatmap(table: pd.DataFrame, value_col: str, title: str, output_path: Path) -> None:
    pivot = table.pivot_table(index="short_window", columns="long_window", values=value_col, aggfunc="mean")
    fig, axis = plt.subplots(figsize=(8, 6))
    image = axis.imshow(pivot.values, aspect="auto", cmap="viridis")
    axis.set_xticks(range(len(pivot.columns)), labels=[str(value) for value in pivot.columns])
    axis.set_yticks(range(len(pivot.index)), labels=[str(value) for value in pivot.index])
    axis.set_xlabel("Long SMA")
    axis.set_ylabel("Short SMA")
    axis.set_title(title)
    fig.colorbar(image, ax=axis)
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_train_test(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(8, 5))
    table.plot(kind="bar", x="label", y=["cagr", "sharpe"], ax=axis)
    axis.set_title("Train vs test performance")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_multi_asset(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(11, 6))
    table.plot(kind="bar", x="ticker", y=["cagr", "benchmark_cagr"], ax=axis)
    axis.set_title("Multi-asset CAGR comparison")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_leaderboard(table: pd.DataFrame, output_path: Path) -> None:
    fig, axis = plt.subplots(figsize=(11, 6))
    top = table.head(10)
    top.plot(kind="bar", x="variant", y=["sharpe", "cagr"], ax=axis)
    axis.set_title("Top model variants")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v03_equity_drawdown(curve: pd.DataFrame, output_path: Path) -> None:
    if curve.empty:
        return
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    axes[0].plot(curve.index, curve["strategy_equity"], label="selected v3")
    axes[0].plot(curve.index, curve["buy_hold_equity"], label="AAPL buy and hold")
    axes[0].set_title("v0.3 selected model equity")
    axes[0].grid(alpha=0.25)
    axes[0].legend()
    axes[1].plot(curve.index, curve["strategy_drawdown"], label="selected v3 drawdown")
    axes[1].plot(curve.index, curve["buy_hold_drawdown"], label="AAPL drawdown")
    axes[1].set_title("v0.3 selected model drawdown")
    axes[1].grid(alpha=0.25)
    axes[1].legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_turnover_scatter(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty:
        return
    fig, axis = plt.subplots(figsize=(10, 6))
    for source, group in table.groupby("source"):
        axis.scatter(group["turnover"], group["sharpe"], label=source, alpha=0.75)
    axis.set_title("Turnover vs Sharpe")
    axis.set_xlabel("Annualized turnover")
    axis.set_ylabel("Sharpe")
    axis.grid(alpha=0.25)
    axis.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_capture(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty:
        return
    fig, axis = plt.subplots(figsize=(10, 6))
    table.plot(kind="bar", x="model", y=["upside_capture", "downside_capture"], ax=axis)
    axis.set_title("Capture ratios vs AAPL")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_allocation_exposure(curve: pd.DataFrame, output_path: Path) -> None:
    if curve.empty or "position" not in curve.columns:
        return
    fig, axis = plt.subplots(figsize=(12, 4))
    axis.plot(curve.index, curve["position"], label="AAPL exposure")
    if "fallback_position" in curve.columns:
        axis.plot(curve.index, curve["fallback_position"], label="fallback exposure")
    axis.set_title("Selected model exposure")
    axis.set_ylim(-0.05, 1.05)
    axis.grid(alpha=0.25)
    axis.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v03_cost_sensitivity(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty:
        return
    ordered = table.sort_values("cost_bps")
    fig, axis = plt.subplots(figsize=(9, 5))
    axis.plot(ordered["cost_bps"], ordered["cagr"], marker="o", label="CAGR")
    axis.plot(ordered["cost_bps"], ordered["sharpe"], marker="s", linestyle="--", label="Sharpe")
    axis.set_title("v0.3 selected model cost sensitivity")
    axis.set_xlabel("Cost assumption, bps")
    axis.grid(alpha=0.25)
    axis.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v03_entry_exit_signals(curve: pd.DataFrame, output_path: Path) -> None:
    if curve.empty or "price" not in curve.columns or "position" not in curve.columns:
        return

    position_change = curve["position"].diff().fillna(curve["position"])
    entries = curve[position_change > 0]
    exits = curve[position_change < 0]

    fig, axes = plt.subplots(2, 1, figsize=(13, 8), sharex=True, height_ratios=[3, 1])
    axes[0].plot(curve.index, curve["price"], label="AAPL price", color="#111827", linewidth=1.4)
    if "short_sma" in curve.columns:
        axes[0].plot(curve.index, curve["short_sma"], label="Short SMA", color="#2563EB", linewidth=1.0, alpha=0.85)
    if "long_sma" in curve.columns:
        axes[0].plot(curve.index, curve["long_sma"], label="Long SMA", color="#F59E0B", linewidth=1.0, alpha=0.85)
    axes[0].scatter(entries.index, entries["price"], marker="^", s=80, color="#16A34A", label="Entry", zorder=5)
    axes[0].scatter(exits.index, exits["price"], marker="v", s=80, color="#DC2626", label="Exit", zorder=5)
    axes[0].set_title("Selected v0.3 model: AAPL entries and exits")
    axes[0].set_ylabel("Adjusted price")
    axes[0].grid(alpha=0.25)
    axes[0].legend(loc="upper left", ncols=3)

    axes[1].step(curve.index, curve["position"], where="post", color="#2563EB", label="AAPL exposure")
    if "fallback_position" in curve.columns and curve["fallback_position"].abs().sum() > 0:
        axes[1].step(curve.index, curve["fallback_position"], where="post", color="#7C3AED", label="Fallback exposure")
    axes[1].set_ylabel("Weight")
    axes[1].set_ylim(-0.05, 1.05)
    axes[1].grid(alpha=0.25)
    axes[1].legend(loc="upper left")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v04_equity_drawdown(curve: pd.DataFrame, output_path: Path) -> None:
    if curve.empty:
        return
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    axes[0].plot(curve.index, curve["strategy_equity"], label="selected v0.4 / retained baseline")
    axes[0].plot(curve.index, curve["buy_hold_equity"], label="AAPL buy and hold")
    axes[0].set_title("v0.4 selected model equity")
    axes[0].grid(alpha=0.25)
    axes[0].legend()
    axes[1].plot(curve.index, curve["strategy_drawdown"], label="model drawdown")
    axes[1].plot(curve.index, curve["buy_hold_drawdown"], label="AAPL drawdown")
    axes[1].set_title("v0.4 selected model drawdown")
    axes[1].grid(alpha=0.25)
    axes[1].legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v04_entry_exit_signals(curve: pd.DataFrame, output_path: Path) -> None:
    if curve.empty or "price" not in curve.columns or "position" not in curve.columns:
        return

    position_change = curve["position"].diff().fillna(curve["position"])
    entries = curve[position_change > 0]
    exits = curve[position_change < 0]

    fig, axes = plt.subplots(3, 1, figsize=(13, 9), sharex=True, height_ratios=[3, 1, 1])
    axes[0].plot(curve.index, curve["price"], label="AAPL price", color="#111827", linewidth=1.4)
    if "short_sma" in curve.columns:
        axes[0].plot(curve.index, curve["short_sma"], label="Short SMA", color="#2563EB", linewidth=1.0)
    if "long_sma" in curve.columns:
        axes[0].plot(curve.index, curve["long_sma"], label="Long SMA", color="#F59E0B", linewidth=1.0)
    if "price_sma" in curve.columns:
        axes[0].plot(curve.index, curve["price_sma"], label="Risk SMA", color="#64748B", linewidth=1.0)
    axes[0].scatter(entries.index, entries["price"], marker="^", s=80, color="#16A34A", label="Entry", zorder=5)
    axes[0].scatter(exits.index, exits["price"], marker="v", s=80, color="#DC2626", label="Exit", zorder=5)
    axes[0].set_title("v0.4 model: AAPL entries, exits, and risk filter")
    axes[0].grid(alpha=0.25)
    axes[0].legend(loc="upper left", ncols=4)

    axes[1].step(curve.index, curve["position"], where="post", label="AAPL exposure", color="#2563EB")
    if "fallback_position" in curve.columns:
        axes[1].step(curve.index, curve["fallback_position"], where="post", label="Fallback exposure", color="#7C3AED")
    axes[1].set_ylim(-0.05, 1.05)
    axes[1].set_ylabel("Weight")
    axes[1].grid(alpha=0.25)
    axes[1].legend(loc="upper left")

    if "risk_off" in curve.columns:
        axes[2].fill_between(
            curve.index,
            curve["risk_off"].fillna(False).astype(float),
            step="post",
            color="#DC2626",
            alpha=0.25,
            label="Risk off",
        )
    if "volatility_weight" in curve.columns:
        axes[2].plot(curve.index, curve["volatility_weight"], color="#0F766E", label="Volatility weight")
    axes[2].set_ylim(-0.05, 1.05)
    axes[2].set_ylabel("Risk")
    axes[2].grid(alpha=0.25)
    if axes[2].get_legend_handles_labels()[0]:
        axes[2].legend(loc="upper left")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v04_capture(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty or not {"model", "upside_capture", "downside_capture"}.issubset(table.columns):
        return
    fig, axis = plt.subplots(figsize=(10, 6))
    table.plot(kind="bar", x="model", y=["upside_capture", "downside_capture", "capture_spread"], ax=axis)
    axis.set_title("v0.4 capture profile")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_regime_performance(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty or "regime" not in table.columns:
        return
    fig, axis = plt.subplots(figsize=(10, 6))
    table.plot(kind="bar", x="regime", y=["strategy_return", "benchmark_return"], ax=axis)
    axis.set_title("Performance by market regime")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_turnover_capture_spread(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty or not {"turnover", "capture_spread"}.issubset(table.columns):
        return
    fig, axis = plt.subplots(figsize=(10, 6))
    scatter = axis.scatter(table["turnover"], table["capture_spread"], c=table["sharpe"], cmap="viridis", alpha=0.75)
    axis.set_title("Turnover vs capture spread")
    axis.set_xlabel("Annualized turnover")
    axis.set_ylabel("Upside capture - downside capture")
    axis.grid(alpha=0.25)
    fig.colorbar(scatter, ax=axis, label="Sharpe")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_exposure_sizing(curve: pd.DataFrame, output_path: Path) -> None:
    if curve.empty or "position" not in curve.columns:
        return
    fig, axis = plt.subplots(figsize=(12, 5))
    axis.plot(curve.index, curve["position"], label="AAPL exposure", color="#2563EB")
    if "fallback_position" in curve.columns:
        axis.plot(curve.index, curve["fallback_position"], label="Fallback exposure", color="#7C3AED")
    if "volatility_weight" in curve.columns:
        axis.plot(curve.index, curve["volatility_weight"], label="Volatility cap", color="#0F766E", alpha=0.8)
    axis.set_title("v0.4 exposure sizing")
    axis.set_ylim(-0.05, 1.05)
    axis.grid(alpha=0.25)
    axis.legend()
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_final_walk_forward(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty or not {"model", "window_id", "cagr"}.issubset(table.columns):
        return
    pivot = table.pivot_table(index="window_id", columns="model", values="cagr", aggfunc="mean")
    fig, axis = plt.subplots(figsize=(11, 6))
    pivot.plot(kind="bar", ax=axis)
    axis.set_title("Final models: CAGR per walk-forward window")
    axis.set_xlabel("Window")
    axis.set_ylabel("CAGR")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _plot_v06_comparison(table: pd.DataFrame, output_path: Path) -> None:
    if table.empty or not {"model", "cagr", "sharpe"}.issubset(table.columns):
        return
    fig, axis = plt.subplots(figsize=(11, 6))
    table.plot(kind="bar", x="model", y=["cagr", "sharpe"], ax=axis)
    axis.set_title("v0.6 candidates on the test period")
    axis.grid(alpha=0.25, axis="y")
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)


def _write_dashboard(sheet, result: ResearchResult) -> None:
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "AAPL Trend Allocation Research Framework"
    sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="0F172A")
    sheet.merge_cells("A1:H1")

    baseline = result.base_backtest.iloc[0]
    best = result.model_leaderboard.iloc[0]
    selected = result.v03_comparison[result.v03_comparison["model"] == "selected_v3"].iloc[0]
    summary = [
        ["Baseline CAGR", baseline["cagr"]],
        ["Baseline Sharpe", baseline["sharpe"]],
        ["Baseline Max Drawdown", baseline["max_drawdown"]],
        ["Best Variant", best["variant"]],
        ["Best Variant CAGR", best["cagr"]],
        ["Best Variant Sharpe", best["sharpe"]],
        ["Selected v0.3 Variant", selected["variant"]],
        ["Selected v0.3 CAGR", selected["cagr"]],
        ["Selected v0.3 Turnover", selected["turnover"]],
    ]
    for row_idx, row in enumerate(summary, start=3):
        sheet.cell(row_idx, 1, row[0])
        sheet.cell(row_idx, 2, row[1])
    _write_table(sheet, result.model_leaderboard.head(12), start_row=11, table_name="DashboardLeaderboard")
    _add_dashboard_chart(sheet, result)
    _set_widths(sheet, [24, 18, 16, 16, 16, 16, 16, 16])


def _add_dashboard_chart(sheet, result: ResearchResult) -> None:
    start_row = 26
    table = result.cost_sensitivity[["label", "cost_bps", "cagr"]].copy()
    table["scenario"] = table["label"] + " " + table["cost_bps"].astype(str) + "bps"
    compact = table[["scenario", "cagr"]]
    _write_table(sheet, compact, start_row=start_row, table_name="DashboardCosts")
    chart = BarChart()
    chart.title = "Cost sensitivity CAGR"
    data = Reference(sheet, min_col=2, min_row=start_row, max_row=start_row + len(compact))
    cats = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(compact))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    sheet.add_chart(chart, "D26")


def _write_table(sheet, table: pd.DataFrame, start_row: int = 1, table_name: str | None = None) -> None:
    if table.empty:
        sheet.cell(start_row, 1, "No rows")
        return
    values = [list(table.columns)] + table.replace({pd.NA: None}).where(pd.notna(table), None).values.tolist()
    for row_offset, row in enumerate(values):
        for col_offset, value in enumerate(row):
            cell = sheet.cell(start_row + row_offset, 1 + col_offset, value)
            if row_offset == 0:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="111827")
    table_ref = (
        f"A{start_row}:"
        f"{get_column_letter(len(table.columns))}{start_row + len(table)}"
    )
    name = table_name or sheet.title.replace(" ", "") + "Table"
    excel_table = Table(displayName=_safe_table_name(name), ref=table_ref)
    excel_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(excel_table)
    _set_widths(sheet, [max(12, min(28, len(str(col)) + 4)) for col in table.columns])


def _raw_results(result: ResearchResult) -> pd.DataFrame:
    return pd.concat(
        [
            result.cost_sensitivity.assign(source="cost_sensitivity"),
            result.parameter_sweep.assign(source="parameter_sweep"),
            result.train_test_results.assign(source="train_test"),
            result.walk_forward_results.assign(source="walk_forward"),
            result.multi_asset_results.assign(source="multi_asset"),
            result.model_leaderboard.assign(source="leaderboard"),
            result.hysteresis_sweep.assign(source="hysteresis_sweep"),
            result.allocation_leaderboard.assign(source="allocation_leaderboard"),
            result.capture_analysis.assign(source="capture_analysis"),
            result.turnover_analysis.assign(source="turnover_analysis"),
            result.v03_comparison.assign(source="v03_comparison"),
            result.v03_cost_sensitivity.assign(source="v03_cost_sensitivity"),
            result.capture_leaderboard.assign(source="capture_leaderboard"),
            result.risk_filter_sweep.assign(source="risk_filter_sweep"),
            result.regime_results.assign(source="regime_results"),
            result.trade_log.assign(source="trade_log"),
            result.benchmark_comparison.assign(source="benchmark_comparison"),
            result.v04_comparison.assign(source="v04_comparison"),
            result.v04_cost_sensitivity.assign(source="v04_cost_sensitivity"),
            result.final_walk_forward.assign(source="final_walk_forward"),
            result.significance_results.assign(source="significance_results"),
            result.family_leaderboard.assign(source="family_leaderboard"),
            result.ensemble_leaderboard.assign(source="ensemble_leaderboard"),
            result.overlay_leaderboard.assign(source="overlay_leaderboard"),
            result.v06_comparison.assign(source="v06_comparison"),
            result.nested_ensemble_walk_forward.assign(source="nested_ensemble_walk_forward"),
        ],
        ignore_index=True,
        sort=False,
    )


def _set_widths(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _safe_table_name(name: str) -> str:
    return "".join(char for char in name if char.isalnum() or char == "_")[:240]
